"""
Parses the Hugessen Compensation Benchmarking Excel (Raw Data + Peer Group tabs).

Column positions and even some header labels (e.g. the Peer Group ticker column,
which is stamped "FY2024"/"FY2025" rather than "Ticker") shift between engagement
years, so columns are identified by header text search wherever the text is unique.
Two pairs of columns share literally identical header text after whitespace
normalization ("Annualized / Converted" appears once for base salary and once for
STIP actual) — those are anchored positionally, relative to the distinctive header
immediately preceding them, instead of by name.
"""
import io
import re
from typing import Any, List, Optional, Tuple

import openpyxl

from models.benchmarking import BenchmarkData, PeerRecord, PeerSizeRecord, PriorYearRecord

_NULL_TOKENS = {"", "-", "–", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "N/A", "(INVALID IDENTIFIER)"}
_ROLE_MATCH_RE = re.compile(r"^Role Match(?:\s+(\d+))?$")


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _is_null(v: Any) -> bool:
    if v is None:
        return True
    return _norm(v).upper() in {t.upper() for t in _NULL_TOKENS}


def _num(v: Any) -> Optional[float]:
    if _is_null(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str(v: Any) -> Optional[str]:
    if _is_null(v):
        return None
    return _norm(v)


def _find_label_value(ws, label: str, max_row: int = 20, max_col: int = 16) -> Optional[Any]:
    """Scan the top-left INPUTS block for a cell matching `label`; return the cell to its
    right. The INPUTS block's column position shifts between template versions (a 2026
    template has an extra Look-Up column pushing everything right by one), so this must
    be a text search, not a fixed coordinate."""
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row), max_col=min(max_col, ws.max_column)):
        for c in row:
            if _norm(c.value).lower() == label.lower():
                return ws.cell(row=c.row, column=c.column + 1).value
    return None


def _find_header_row(ws, anchor: str, max_row: int = 100) -> int:
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        for c in row:
            if _norm(c.value) == anchor:
                return c.row
    raise ValueError(f'Could not locate header row (looking for a "{anchor}" column).')


def _detect_role_match_columns(headers: List[str]) -> List[int]:
    """0-based indices of every Role Match / Role Match N column, sorted numerically.
    No fixed cap — a template adding "Role Match 3" or "Role Match 4" works without a
    code change."""
    found: List[Tuple[int, int]] = []
    for i, h in enumerate(headers):
        m = _ROLE_MATCH_RE.match(h)
        if m:
            n = int(m.group(1)) if m.group(1) else 1
            found.append((n, i))
    found.sort(key=lambda t: t[0])
    return [i for _, i in found]


def parse_benchmarking_excel(file_bytes: bytes) -> BenchmarkData:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)

    if "Raw Data" not in wb.sheetnames:
        raise ValueError('Workbook is missing a "Raw Data" tab.')
    if "Peer Group" not in wb.sheetnames:
        raise ValueError('Workbook is missing a "Peer Group" tab.')

    warnings: List[str] = []

    # ---- Raw Data: INPUTS block ---------------------------------------------
    raw_ws = wb["Raw Data"]
    client_ticker = _str(_find_label_value(raw_ws, "Client Ticker")) or ""
    reporting_currency = _str(_find_label_value(raw_ws, "Currency")) or "CAD"
    if not client_ticker:
        warnings.append('Could not find "Client Ticker" in the Raw Data inputs section.')

    # ---- Raw Data: header row + column map ----------------------------------
    header_row_idx = _find_header_row(raw_ws, "Look-Up")
    headers = [_norm(c.value) for c in next(raw_ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))]

    def _col(name: str) -> Optional[int]:
        try:
            return headers.index(name)
        except ValueError:
            return None

    def _col_after(anchor_name: str) -> Optional[int]:
        i = _col(anchor_name)
        return i + 1 if i is not None else None

    idx = {
        "lookup": _col("Look-Up"),
        "ticker": _col("Ticker"),
        "company": _col("Company"),
        "year": _col("Year"),
        "position_match": _col("Position Match"),
        "position_title": _col("Position Title"),
        "incumbent": _col("Incumbent"),
        "base_salary": _col_after("Actual Paid"),
        "base_salary_next": _col("Latest (if disclosed) - Annualized / Converted"),
        "stip_actual": _col_after("Actual $"),
        "stip_target_pct": _col("Target STI %"),
        "stip_3yr_pct": _col("3 Year Average STIP %"),
        "ltip_3yr_pct": _col("3 Year Average LTIP %"),
        "stip_target_dollar": _col("Target STI $"),
        "actual_tcc": _col("Actual TCC"),
        "target_tcc": _col("Target TCC"),
        "rsu": _col("RSUs"),
        "psu": _col("PSUs"),
        "options": _col("Stock Options"),
        "dsu": _col("DSUs"),
        "lt_cash": _col("LT Cash"),
        "ltip_total": _col("Total LTI $"),
        "ltip_target_pct": _col("Target LTI %"),
        "ltip_target_dollar": _col("Target LTI $"),
        "pension": _col("Pension"),
        "pension_type": _col("DB or DC Pension?"),
        "pension_converted": _col("Pension Converted"),
        "pension_3yr": _col("Pension 3 Year Average"),
        "other": _col("Other"),
        "other_converted": _col("Other Converted"),
        "other_3yr": _col("Other 3 Year Average"),
        "months": _col("Months"),
    }

    # "Actual (base + actual bonus + actual LTIP)" vs "Actual (Actual TDC + Pension + Other)"
    # — both start with "Actual (" after whitespace normalization, disambiguated by content.
    idx["actual_tdc"] = idx["target_tdc"] = idx["actual_total_comp"] = idx["target_total_comp"] = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if hl.startswith("actual (") and "pension" in hl:
            idx["actual_total_comp"] = i
        elif hl.startswith("actual (") and "ltip" in hl:
            idx["actual_tdc"] = i
        elif hl.startswith("target (") and "pension" in hl:
            idx["target_total_comp"] = i
        elif hl.startswith("target (") and "ltip" in hl:
            idx["target_tdc"] = i

    role_match_cols = _detect_role_match_columns(headers)
    if not role_match_cols:
        warnings.append('No "Role Match" column detected — role tabs will be built from Position Match only.')

    required = {"lookup": "Look-Up", "ticker": "Ticker", "company": "Company", "year": "Year"}
    missing = [label for key, label in required.items() if idx[key] is None]
    if missing:
        raise ValueError(f"Raw Data tab is missing required column(s): {', '.join(missing)}")

    def cell(values: List[Any], key: str) -> Any:
        i = idx[key]
        return values[i] if i is not None and i < len(values) else None

    def rolecell(values: List[Any], col: int) -> Any:
        return values[col] if col < len(values) else None

    peers: List[PeerRecord] = []
    years_seen = set()

    for row in raw_ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        lookup_key = _str(cell(row, "lookup"))
        if lookup_key is None:
            continue

        year_raw = cell(row, "year")
        if _is_null(year_raw):
            continue
        try:
            year = int(year_raw)
        except (TypeError, ValueError):
            continue

        company_name = _str(cell(row, "company"))
        if company_name is None:
            # Template buffer rows below the real peer roster still evaluate their
            # Look-Up formula to a non-blank string (e.g. "0_2025_CEO"), but the
            # Company cell is blank or a CIQ "(Invalid Identifier)" error — these
            # are not real data rows.
            continue

        years_seen.add(year)
        position_match = _str(cell(row, "position_match"))
        role_values = [v for v in (_str(rolecell(row, c)) for c in role_match_cols) if v]
        # Position Match and Role Match are additive, not override: every row keeps
        # its baseline position label (CEO/CFO/NEO1/NEO2/NEO3...) AND gains any Role
        # Match labels on top, so a row can appear under multiple tabs at once
        # (e.g. NEO1 + COO + "2nd Highest" are three simultaneous, valid tags for
        # the same person). Every benchmarked company has one row per position slot,
        # so each position tab's count should match the company count.
        role_tags = list(dict.fromkeys(([position_match] if position_match else []) + role_values))
        if not role_tags:
            warnings.append(f'Row "{lookup_key}" has no usable Role Match or Position Match value — excluded from role tabs.')

        ticker = _str(cell(row, "ticker")) or ""

        peers.append(PeerRecord(
            lookup_key=lookup_key,
            ticker=ticker,
            company_name=company_name,
            year=year,
            role_tags=role_tags,
            position_match=position_match,
            position_title=_str(cell(row, "position_title")),
            incumbent_name=_str(cell(row, "incumbent")),
            base_salary=_num(cell(row, "base_salary")),
            base_salary_next_year=_num(cell(row, "base_salary_next")),
            stip_actual=_num(cell(row, "stip_actual")),
            stip_target_pct=_num(cell(row, "stip_target_pct")),
            stip_3yr_avg_pct=_num(cell(row, "stip_3yr_pct")),
            ltip_3yr_avg_pct=_num(cell(row, "ltip_3yr_pct")),
            stip_target_dollar=_num(cell(row, "stip_target_dollar")),
            actual_tcc=_num(cell(row, "actual_tcc")),
            target_tcc=_num(cell(row, "target_tcc")),
            ltip_rsu=_num(cell(row, "rsu")),
            ltip_psu=_num(cell(row, "psu")),
            ltip_options=_num(cell(row, "options")),
            ltip_dsu=_num(cell(row, "dsu")),
            ltip_cash=_num(cell(row, "lt_cash")),
            ltip_total=_num(cell(row, "ltip_total")),
            ltip_target_pct=_num(cell(row, "ltip_target_pct")),
            ltip_target_dollar=_num(cell(row, "ltip_target_dollar")),
            target_tdc=_num(cell(row, "target_tdc")),
            actual_tdc=_num(cell(row, "actual_tdc")),
            pension=_num(cell(row, "pension")),
            pension_type=_str(cell(row, "pension_type")),
            pension_converted=_num(cell(row, "pension_converted")),
            pension_3yr_avg=_num(cell(row, "pension_3yr")),
            other=_num(cell(row, "other")),
            other_converted=_num(cell(row, "other_converted")),
            other_3yr_avg=_num(cell(row, "other_3yr")),
            actual_total_comp=_num(cell(row, "actual_total_comp")),
            target_total_comp=_num(cell(row, "target_total_comp")),
            months_in_role=_num(cell(row, "months")),
            is_client=(ticker == client_ticker),
        ))

    if not peers:
        raise ValueError("No data rows found below the Raw Data header row.")

    # ---- Peer Group tab -------------------------------------------------------
    peer_ws = wb["Peer Group"]
    pg_header_row = _find_header_row(peer_ws, "Company Name")
    pg_headers = [_norm(c.value) for c in next(peer_ws.iter_rows(min_row=pg_header_row, max_row=pg_header_row))]

    def _pg_col(*needles: str) -> Optional[int]:
        for i, h in enumerate(pg_headers):
            hl = h.lower()
            if all(n in hl for n in needles):
                return i
        return None

    pg_idx_company = _pg_col("company", "name")
    pg_idx_tev = _pg_col("enterprise", "value")
    pg_idx_mktcap = _pg_col("market", "cap")
    pg_idx_assets = _pg_col("total", "assets")
    pg_idx_revenue = _pg_col("revenue")

    def _unit_multiplier(col: Optional[int]) -> float:
        # Peer Group size columns are typically labeled "($MM)" / "($B)" — values
        # must be scaled to raw currency units so they're comparable in magnitude
        # to pay values (base salary, TCC, etc.), which are never in millions.
        if col is None:
            return 1.0
        h = pg_headers[col].lower()
        if "$mm" in h or "millions" in h:
            return 1_000_000.0
        if "$b" in h or "billions" in h:
            return 1_000_000_000.0
        return 1.0

    peer_sizes: List[PeerSizeRecord] = []
    for row in peer_ws.iter_rows(min_row=pg_header_row + 1, values_only=True):
        # The ticker column's header is a dynamic year stamp (e.g. "FY2024"), not
        # literally "Ticker" — it cannot be found by name search, so it's anchored
        # positionally as the leftmost column instead.
        ticker = _str(row[0]) if row else None
        company_name = _str(row[pg_idx_company]) if pg_idx_company is not None and pg_idx_company < len(row) else None
        if not ticker or not company_name:
            continue

        def pg_num(col: Optional[int], field: str) -> Optional[float]:
            if col is None or col >= len(row):
                return None
            v = row[col]
            n = _num(v)
            if v is not None and n is None:
                warnings.append(f'Peer Group row "{company_name}": "{field}" is missing or formula-errored — stored as null.')
                return None
            return n * _unit_multiplier(col) if n is not None else None

        peer_sizes.append(PeerSizeRecord(
            ticker=ticker,
            company_name=company_name,
            market_cap=pg_num(pg_idx_mktcap, "Market Cap"),
            tev=pg_num(pg_idx_tev, "TEV"),
            revenue=pg_num(pg_idx_revenue, "Revenue"),
            total_assets=pg_num(pg_idx_assets, "Total Assets"),
        ))

    roles = sorted({tag for p in peers for tag in p.role_tags})
    available_years = sorted(years_seen)
    default_year = max(available_years) if available_years else 0

    # ---- Aging factor ("Data Aging:" label in the Raw Data inputs block) ------
    aging_factor = _num(_find_label_value(raw_ws, "Data Aging:", max_row=30))
    if aging_factor is None:
        aging_factor = 0.03
        warnings.append('Could not find a "Data Aging:" value in the Raw Data inputs — defaulting to 3%.')

    # ---- Optional "Import <year> Data" tab (prior engagement's final TDC) -----
    prior_year_records: List[PriorYearRecord] = []
    import_re = re.compile(r"^Import (\d{4}) Data$")
    for sheet_name in wb.sheetnames:
        m = import_re.match(sheet_name.strip())
        if not m:
            continue
        prior_year_records.extend(
            _parse_import_sheet(wb[sheet_name], int(m.group(1)), client_ticker, warnings)
        )

    return BenchmarkData(
        client_ticker=client_ticker,
        reporting_currency=reporting_currency,
        default_year=default_year,
        available_years=available_years,
        roles=roles,
        peers=peers,
        peer_sizes=peer_sizes,
        validation_warnings=warnings,
        aging_factor=aging_factor,
        prior_year_records=prior_year_records,
    )


def _parse_import_sheet(ws, year: int, client_ticker: str, warnings: List[str]) -> List[PriorYearRecord]:
    """Parse an "Import <year> Data" tab: Lookup / Role match / Company /
    Actual TDC / Target TDC.

    The sheet mixes units — client rows are pasted in raw dollars while peer
    rows carry the prior model's $000s. Values under the threshold are scaled
    up so everything is stored in raw dollars like the Raw Data pay fields.
    """
    THOUSANDS_THRESHOLD = 100_000.0

    header_row = None
    cols: dict = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        labels = {_norm(c.value).lower(): c.column - 1 for c in row if c.value is not None}
        if "lookup" in labels and "target tdc" in labels:
            header_row = row[0].row
            cols = {
                "lookup": labels.get("lookup"),
                "role": labels.get("role match"),
                "company": labels.get("company"),
                "actual_tdc": labels.get("actual tdc"),
                "target_tdc": labels.get("target tdc"),
            }
            break
    if header_row is None:
        warnings.append(f'"{ws.title}": could not locate Lookup/Target TDC headers — prior-year data skipped.')
        return []

    def scaled(v) -> Optional[float]:
        n = _num(v)
        if n is None:
            return None
        return n * 1000.0 if 0 < n < THOUSANDS_THRESHOLD else n

    records: List[PriorYearRecord] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def at(key):
            i = cols.get(key)
            return row[i] if i is not None and i < len(row) else None

        lookup = _str(at("lookup"))
        role = _str(at("role"))
        company = _str(at("company"))
        if not lookup or not company or not role:
            continue
        ticker = lookup.split("_")[0] if "_" in lookup else ""
        records.append(PriorYearRecord(
            lookup_key=lookup,
            ticker=ticker,
            role=role,
            company_name=company,
            year=year,
            actual_tdc=scaled(at("actual_tdc")),
            target_tdc=scaled(at("target_tdc")),
            is_client=(ticker == client_ticker),
        ))
    return records
