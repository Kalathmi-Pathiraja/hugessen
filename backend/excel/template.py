"""
Generate the LTIP Excel data-entry template (openpyxl).
User fills this in and uploads it for market data ingestion.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


NAVY = "0A192F"
ORANGE = "FF6B00"
LIGHT_GREY = "E2E8F0"
OFF_WHITE = "F8FAFC"
WHITE = "FFFFFF"


def _header_style(ws, row, col, value, bg=NAVY, font_color=WHITE, bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def _note_style(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(italic=True, color="64748B", name="Calibri", size=10)
    return cell


def _thin_border():
    thin = Side(style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build_ltip_template() -> bytes:
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Price History
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Price History"

    ws1.merge_cells("A1:F1")
    banner = ws1["A1"]
    banner.value = "LTIP Market Data — Price History"
    banner.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    banner.fill = PatternFill("solid", fgColor=NAVY)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    _note_style(ws1, 2, 1,
        "Instructions: Enter one row per trading date. "
        "Date format: YYYY-MM-DD. ~756 rows for 3 years of data. "
        "Column B header must be exactly 'COMPANY'. "
        "Add one column per peer with their ticker as the header.")
    ws1.merge_cells("A2:F2")

    headers = ["Date", "COMPANY", "PEER_1", "PEER_2", "PEER_3", "PEER_4"]
    for col, h in enumerate(headers, 1):
        _header_style(ws1, 3, col, h, bg=ORANGE if h == "COMPANY" else NAVY)

    # Sample data rows
    sample_dates = ["2022-01-03", "2022-01-04", "2022-01-05", "...", "(continue to ~756 rows)"]
    for row_i, dt in enumerate(sample_dates, 4):
        ws1.cell(row=row_i, column=1, value=dt).font = Font(name="Calibri", size=10)
        for col in range(2, 7):
            cell = ws1.cell(row=row_i, column=col, value="" if dt.startswith(".") else None)
            cell.number_format = "0.00"

    ws1.column_dimensions["A"].width = 14
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col_letter].width = 12

    # -----------------------------------------------------------------------
    # Sheet 2: Peer Config
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Peer Config")

    ws2.merge_cells("A1:F1")
    banner2 = ws2["A1"]
    banner2.value = "Peer Configuration"
    banner2.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    banner2.fill = PatternFill("solid", fgColor=NAVY)
    banner2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _note_style(ws2, 2, 1,
        "Ticker must exactly match the column header in the Price History sheet. "
        "Leave Dividend Yield blank if price data already includes reinvested dividends. "
        "Beta (optional): 2-year weekly or 5-year monthly regression beta vs. relevant index "
        "(source: Bloomberg, Capital IQ, or Yahoo Finance). Leave blank to use common drift.")
    ws2.merge_cells("A2:F2")

    peer_headers = [
        "Ticker", "Company Name", "Peer Type\n(Custom / Index)",
        "Include in TSR Calc\n(Yes / No)", "Dividend Yield (%)\n(e.g. 0.035 = 3.5%)",
        "Beta\n(optional — for CAPM drift)",
    ]
    for col, h in enumerate(peer_headers, 1):
        c = _header_style(ws2, 3, col, h)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[3].height = 42

    # Beta column gets a distinct light teal fill to signal it's optional
    TEAL_LIGHT = "E6F4F1"
    beta_header = ws2.cell(row=3, column=6)
    beta_header.fill = PatternFill("solid", fgColor="0D7A6B")
    beta_header.font = Font(bold=True, color=WHITE, name="Calibri", size=11)

    # Data validation dropdowns
    dv_type = DataValidation(type="list", formula1='"Custom,Index"', allow_blank=True)
    dv_include = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws2.add_data_validation(dv_type)
    ws2.add_data_validation(dv_include)

    sample_peers = [
        ["COMPANY", "Subject Company", "Custom", "Yes", 0.030, 1.10],
        ["PEER_1",  "Peer Company A",  "Custom", "Yes", 0.025, 0.85],
        ["PEER_2",  "Peer Company B",  "Custom", "Yes", 0.000, 1.45],
        ["PEER_3",  "Peer Company C",  "Index",  "Yes", 0.018, None],
    ]
    for row_i, peer in enumerate(sample_peers, 4):
        for col_i, val in enumerate(peer, 1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = _thin_border()
            if col_i == 3:
                dv_type.sqref += cell.coordinate  # type: ignore[operator]
            if col_i == 4:
                dv_include.sqref += cell.coordinate  # type: ignore[operator]
            if col_i == 6:
                cell.fill = PatternFill("solid", fgColor=TEAL_LIGHT)
                if val is not None:
                    cell.number_format = "0.00"

    # Beta note row
    note_row = len(sample_peers) + 4 + 1
    note = ws2.cell(row=note_row, column=6,
                    value="Leave blank → peer uses company growth rate as drift")
    note.font = Font(italic=True, color="64748B", name="Calibri", size=9)

    col_widths = [12, 22, 18, 22, 22, 24]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # -----------------------------------------------------------------------
    # Sheet 3: Assumptions Override
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Assumptions Override")

    ws3.merge_cells("A1:C1")
    banner3 = ws3["A1"]
    banner3.value = "Simulation Assumptions (override defaults below)"
    banner3.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    banner3.fill = PatternFill("solid", fgColor=NAVY)
    banner3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    _header_style(ws3, 2, 1, "Parameter", bg=ORANGE)
    _header_style(ws3, 2, 2, "Default Value", bg=ORANGE)
    _header_style(ws3, 2, 3, "Description", bg=ORANGE)

    assumptions = [
        ("Risk-Free Rate (%)", 3.5,
         "Use current Government of Canada 5-yr bond yield (decimal, e.g. 3.5 = 3.5%)"),
        ("Dividend Yield (%) — Company", 0.0,
         "Annual continuous dividend yield for Black-Scholes. 0 if included in price data."),
        ("Vesting Term (Years)", 3,
         "Performance period length. Drives T in GBM terminal draw."),
        ("Option Term (Years)", 5,
         "Time to expiry for Black-Scholes (typically longer than vesting term)."),
        ("Simulation Count", 10000,
         "Number of Monte Carlo trials. Do not reduce below 5,000."),
    ]

    for row_i, (param, default, desc) in enumerate(assumptions, 3):
        ws3.cell(row=row_i, column=1, value=param).font = Font(bold=True, name="Calibri", size=10)
        ws3.cell(row=row_i, column=2, value=default).font = Font(name="Calibri", size=10)
        ws3.cell(row=row_i, column=3, value=desc).font = Font(
            italic=True, color="64748B", name="Calibri", size=10)
        for col in range(1, 4):
            ws3.cell(row=row_i, column=col).border = _thin_border()

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 60

    # -----------------------------------------------------------------------
    # Save and return bytes
    # -----------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
